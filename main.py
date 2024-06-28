import requests
import json
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from sklearn.preprocessing import OneHotEncoder, StandardScaler
import holidays
import joblib
from tensorflow import keras
from keras.losses import MeanSquaredError
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from openpyxl import load_workbook
from openpyxl.styles import Font

def get_data():
    print("Choose an option:")
    print("1. Enter a date range")
    print("2. Enter a specific date and time")
    choice = input("Enter your choice (1 or 2): ")

    berlin_tz = ZoneInfo("Europe/Berlin")
    utc_tz = ZoneInfo("UTC")

    if choice == '1':
        start_date = input("Enter start date (dd.mm.yyyy): ")
        end_date = input("Enter end date (dd.mm.yyyy): ")
        start_time = datetime.strptime(start_date, "%d.%m.%Y").replace(hour=0, minute=0, second=0, tzinfo=berlin_tz)
        end_time = datetime.strptime(end_date, "%d.%m.%Y").replace(hour=23, minute=59, second=59, tzinfo=berlin_tz)
    elif choice == '2':
        specific_datetime = input("Enter specific date and time (dd.mm.yyyy HH:MM): ")
        start_time = datetime.strptime(specific_datetime, "%d.%m.%Y %H:%M").replace(tzinfo=berlin_tz)
        end_time = start_time + timedelta(minutes=59)  # Set end time to 59 minutes after start time
    else:
        print("Invalid choice. Exiting.")
        return None, None

    start_time_utc = start_time.astimezone(utc_tz)
    end_time_utc = end_time.astimezone(utc_tz)

    start_time_str = start_time_utc.strftime("%Y-%m-%dT%H:%M:%S%z")
    end_time_str = end_time_utc.strftime("%Y-%m-%dT%H:%M:%S%z")

    start_time_str = start_time_str[:-2] + ':' + start_time_str[-2:]
    end_time_str = end_time_str[:-2] + ':' + end_time_str[-2:]

    print(f"Requesting data from {start_time_str} to {end_time_str}")

    url = "https://api.hystreet.com/locations/257"
    querystring = {"from": start_time_str, "to": end_time_str, "resolution": "hour"}
    headers = {
        "Content-Type": "application/json",
        "X-API-Token": "rnHrWEGabgFWfXJyz5gWosav"
    }

    response = requests.request("GET", url, headers=headers, params=querystring)
    data = json.loads(response.text)

    if 'measurements' not in data:
        print("The key 'measurements' is not present in the JSON response.")
        print("API response:", data)
        return None, None

    dataset_columns = ['location', 'time of measurement', 'weekday', 'pedestrians count', 'temperature in ºc', 'weather condition', 'incidents', 'events']
    rows = []
    for measurement in data['measurements']:
        measurement_time = datetime.strptime(measurement['timestamp'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(berlin_tz)
        row = [
            data['name'],
            measurement_time,
            measurement_time.strftime('%A'),
            measurement['pedestrians_count'],
            measurement['temperature'],
            measurement['weather_condition'],
            '',
            ''
        ]
        rows.append(row)

    return pd.DataFrame(rows, columns=dataset_columns), choice

def preprocess_data(data):
    def get_season(month):
        if month in [12, 1, 2]:
            return 'winter'
        elif month in [3, 4, 5]:
            return 'spring'
        elif month in [6, 7, 8]:
            return 'summer'
        elif month in [9, 10, 11]:
            return 'autumn'

    # Ensure 'time of measurement' is timezone-aware
    data['time of measurement'] = pd.to_datetime(data['time of measurement'], utc=True)

    data['hour'] = data['time of measurement'].dt.hour
    data['day'] = data['time of measurement'].dt.day
    data['month'] = data['time of measurement'].dt.month
    data['weekday'] = data['time of measurement'].dt.weekday
    data['season'] = data['month'].apply(get_season)

    encoder = OneHotEncoder(sparse=False)
    encoded_weather = pd.DataFrame(encoder.fit_transform(data[['weather condition']]), 
                                   columns=encoder.get_feature_names_out(['weather condition']))
    encoded_season = pd.DataFrame(encoder.fit_transform(data[['season']]), 
                                  columns=encoder.get_feature_names_out(['season']))

    possible_weather_conditions = ['weather condition_clear-day', 'weather condition_clear-night', 'weather condition_cloudy',
                                   'weather condition_fog', 'weather condition_partly-cloudy-day', 'weather condition_partly-cloudy-night',
                                   'weather condition_rain', 'weather condition_sleet', 'weather condition_snow', 'weather condition_wind']
    for condition in possible_weather_conditions:
        if condition not in encoded_weather.columns:
            encoded_weather[condition] = 0
    
    possible_seasons = ['season_winter', 'season_spring', 'season_summer', 'season_autumn']
    for season in possible_seasons:
        if season not in encoded_season.columns:
            encoded_season[season] = 0

    data = pd.concat([data, encoded_weather, encoded_season], axis=1)
    data.drop(columns=['weather condition', 'season'], inplace=True)
    
    de_holidays = holidays.Germany(state='SH')
    data['is_holiday'] = data['time of measurement'].apply(lambda x: 1 if x in de_holidays else 0)
    data['weekday_code'] = data['weekday'].apply(lambda x: 1 if x < 5 else 0)

    data['holiday_weather'] = data['is_holiday'] * data['weather condition_clear-day']
    data['weekday_hour'] = data['weekday_code'] * data['hour']

    for idx, row in data.iterrows():
        if row['incidents'] == 'Laserausfall':
            similar_time_data = data[(data['hour'] == row['hour']) & (data['weekday'] == row['weekday'])]
            similar_time_data = similar_time_data[
                (similar_time_data['month'].isin([row['month'] - 1, row['month'], row['month'] + 1]))
            ]
            if not similar_time_data.empty:
                mean_value = similar_time_data['pedestrians count'].mean()
                variance = np.random.uniform(-0.15, 0.15)
                adjusted_value = mean_value * (1 + variance)
                data.at[idx, 'pedestrians count'] = adjusted_value

    data['incident_occurred'] = data['incidents'].apply(lambda x: 1 if pd.notna(x) else 0)
    data.drop(columns=['incidents'], inplace=True)

    features_to_scale = data[['pedestrians count', 'temperature in ºc', 'hour', 'day', 'month', 'weekday_code']]
    scaler = StandardScaler()
    features_scaled = scaler.fit_transform(features_to_scale)
    features_scaled_df = pd.DataFrame(features_scaled, columns=[col + '_scaled' for col in features_to_scale.columns])
    data = pd.concat([data, features_scaled_df], axis=1)
    data.drop(columns=("location"), inplace=True)

    return data

def validate_and_predict(data, choice):
    mse = MeanSquaredError()

    features = [
        'temperature in ºc', 'hour', 'day', 'month', 'is_holiday', 'weekday_code',
        'weather condition_clear-day', 'weather condition_clear-night',
        'weather condition_cloudy', 'weather condition_fog',
        'weather condition_partly-cloudy-day', 'weather condition_partly-cloudy-night',
        'weather condition_rain', 'weather condition_snow', 'weather condition_wind',
        'season_autumn', 'season_spring', 'season_summer', 'season_winter',
        'holiday_weather', 'weekday_hour'
    ]
    target = 'pedestrians count'

    additional_columns = ['time of measurement']
    all_columns = additional_columns + features + [target]

    for feature in features:
        if feature not in data.columns:
            data[feature] = 0

    data = data[all_columns]

    scaler = joblib.load('scaler.pkl')
    pt = joblib.load('power_transformer.pkl')
    pca = joblib.load('pca.pkl')
    model = keras.models.load_model('best_pedestrian_prediction_model.h5', custom_objects={'mse': mse})

    X_test = data[scaler.feature_names_in_]
    X_test_scaled = scaler.transform(X_test)
    X_test_pca = pca.transform(X_test_scaled)

    predictions = model.predict(X_test_pca).flatten()
    predictions = pt.inverse_transform(predictions.reshape(-1, 1)).flatten()
    predictions = np.maximum(predictions, 0)
    predictions = np.ceil(predictions).astype(int)

    actual_values = data[target].values
    results = pd.DataFrame({
        'time of measurement': data['time of measurement'],
        'Actual': actual_values,
        'Predicted': predictions
    })

    results['absolute difference'] = abs(results['Predicted'] - results['Actual'])
    results['percentage difference'] = (results['absolute difference'] / results['Actual'].replace(0, np.nan)) * 100
    results['correctness'] = 100 - results['percentage difference']

    results['percentage difference'] = results['percentage difference'].clip(upper=100).fillna(100)
    results['correctness'] = results['correctness'].clip(lower=0, upper=100).fillna(0)

    mae = mean_absolute_error(results['Actual'], results['Predicted'])
    mse = mean_squared_error(results['Actual'], results['Predicted'])
    rmse = np.sqrt(mse)
    mape = np.mean(np.abs((results['Actual'] - results['Predicted']) / results['Actual'].replace(0, np.nan))) * 100
    r_squared = r2_score(results['Actual'], results['Predicted'])

    mape = np.nan_to_num(mape, nan=0, posinf=100, neginf=100)

    if choice == '1':
        # Convert 'time of measurement' to timezone-naive datetime
        data['time of measurement'] = data['time of measurement'].dt.tz_localize(None)
        results['time of measurement'] = results['time of measurement'].dt.tz_localize(None)

        output_path = 'prediction_results.xlsx'
        with pd.ExcelWriter(output_path) as writer:
            data.to_excel(writer, sheet_name='All Columns', index=False)
            results.to_excel(writer, sheet_name='Specific Columns', index=False)
            pd.DataFrame({
                'KPI': ['Mean Absolute Error (MAE)', 'Mean Squared Error (MSE)', 'Root Mean Squared Error (RMSE)', 'Mean Absolute Percentage Error (MAPE)', 'R-squared (R²)'],
                'Value': [mae, mse, rmse, mape, r_squared]
            }).to_excel(writer, sheet_name='KPIs', index=False)

        wb = load_workbook(output_path)

        def add_average_row_and_format(sheet_name):
            ws = wb[sheet_name]
            
            averages = {}
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=2):
                col_letter = col[0].column_letter
                numeric_values = [cell.value for cell in col if isinstance(cell.value, (int, float))]
                if numeric_values:
                    averages[col_letter] = sum(numeric_values) / len(numeric_values)
            
            ws.insert_rows(2)
            ws.insert_rows(3)
            ws.insert_rows(4)
            
            ws['A2'] = 'Durchschnitt:'
            ws['A2'].font = Font(bold=True)
            
            for col_letter, avg_value in averages.items():
                cell = ws[f'{col_letter}2']
                cell.value = avg_value
                cell.font = Font(bold=True)
            
            ws.insert_rows(5)
            ws.insert_rows(6)

        add_average_row_and_format('Specific Columns')

        wb.save(output_path)
        print(f"Results saved to {output_path}")
    else:
        print("Echte Passantendaten:", results["Actual"].values[0])
        print("Prognose:", results["Predicted"].values[0])
        print(f"Prozentuale Abweichung: {results['percentage difference'].values[0]:.2f}%")
        print(f"Absolute Abweichung: {results['absolute difference'].values[0]}")
        print("\nKPIs:")
        print(f"Mean Absolute Error (MAE): {mae:.2f}")
        print(f"Mean Squared Error (MSE): {mse:.2f}")
        print(f"Root Mean Squared Error (RMSE): {rmse:.2f}")
        print(f"Mean Absolute Percentage Error (MAPE): {mape:.2f}%")
        print(f"R-squared (R²): {r_squared:.4f}")

    return results

def main():
    data, choice = get_data()
    if data is not None:
        preprocessed_data = preprocess_data(data)
        results = validate_and_predict(preprocessed_data, choice)
        if choice == '1':
            print("Data processing and prediction completed successfully.")
        else:
            print("Data processing and prediction for the specific date and time completed successfully.")

if __name__ == "__main__":
    main()